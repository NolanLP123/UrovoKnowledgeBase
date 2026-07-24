#!/usr/bin/env python3
"""Convert knowledge-template.xlsx into public/knowledge.json.

Requires Python 3 and openpyxl:
    pip install openpyxl
    python tools/xlsx-to-json.py knowledge-template.xlsx public/knowledge.json
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def split_lines(value):
    if value is None:
        return []
    return [item.strip() for item in str(value).replace(",", "\n").splitlines() if item.strip()]


def main():
    if len(sys.argv) != 3:
        raise SystemExit("Usage: python tools/xlsx-to-json.py knowledge-template.xlsx public/knowledge.json")
    source, destination = map(Path, sys.argv[1:])
    workbook = load_workbook(source, data_only=True)
    sheet = workbook["Knowledge"]
    headers = [cell.value for cell in sheet[1]]
    entries = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        row = dict(zip(headers, values))
        if not row.get("id") or not row.get("title"):
            continue
        parameters = []
        raw_parameters = row.get("parameters_json")
        if raw_parameters:
            parameters = json.loads(raw_parameters)
        entries.append(
            {
                "id": str(row["id"]).strip(),
                "category": row.get("category") or "USDK API",
                "module": row.get("module") or "",
                "title": row.get("title") or "",
                "summary": row.get("summary") or "",
                "signature": row.get("signature") or "",
                "parameters": parameters,
                "returns": row.get("returns") or "",
                "minAndroid": row.get("min_android") or "",
                "models": split_lines(row.get("models")),
                "status": row.get("status") or "Active",
                "keywords": row.get("keywords") or "",
                "source": row.get("source") or "",
                "version": row.get("version") or "",
            }
        )
    categories_sheet = workbook["Categories"]
    categories = []
    for name, enabled, description in categories_sheet.iter_rows(min_row=2, values_only=True):
        if name:
            categories.append(
                {"name": name, "enabled": str(enabled).strip().lower() in {"yes", "true", "1", "是"}, "description": description or ""}
            )
    payload = {
        "meta": {
            "title": "Urovo Technical Knowledge Base",
            "subtitle": "USDK API and enterprise device management reference",
            "version": entries[0]["version"] if entries else "",
            "entryCount": len(entries),
            "generatedFrom": source.name,
        },
        "categories": categories,
        "entries": entries,
    }
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Generated {destination} with {len(entries)} entries")


if __name__ == "__main__":
    main()
